from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from database import (
    SessionLocal, Camera, Course, CameraCommand, Attendance, Enrollment,
    Session, get_batch_semester_map, normalize_semester
)
from functools import wraps
from datetime import datetime
from routes.admin import roll_batch_sort_key

cr_bp = Blueprint("cr", __name__, url_prefix="/cr")
def cr_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if current_user.role != "cr":
            flash("এই পেজ দেখার অনুমতি আপনার নেই।", "error")
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated


def get_current_cr_semester(db):
    """
    CR এর batch (current_user.session_id) এই মুহূর্তে কোন semester এ আছে,
    সেটা Excel data থেকে বের করে। কোনো ম্যাচ না পেলে None রিটার্ন করে।
    """
    cr_session = db.query(Session).get(current_user.session_id)
    if not cr_session:
        return None

    normalized_semester_to_batch = get_batch_semester_map()
    for norm_sem, batch_name in normalized_semester_to_batch.items():
        if batch_name == cr_session.name:
            return norm_sem 
    return None


@cr_bp.route("/dashboard", methods=["GET", "POST"])
@login_required
@cr_required
def dashboard():
    db = SessionLocal()
    try:
        if request.method == "POST":
            camera_id = request.form.get("camera_id")
            course_id = request.form.get("course_id")

            if not camera_id or not course_id:
                flash("Camera এবং Course দুইটাই সিলেক্ট করুন!", "error")
            else:
                new_command = CameraCommand(
                    camera_id=camera_id,
                    course_id=course_id,
                    session_id=current_user.session_id,
                    status="pending"
                )
                db.add(new_command)

                camera_obj = db.query(Camera).get(camera_id)
                if camera_obj:
                    camera_obj.current_course_id = course_id
                    camera_obj.current_session_id = current_user.session_id

                db.commit()
                flash("Camera command পাঠানো হয়েছে!", "success")

        cameras = db.query(Camera).order_by(Camera.camera_code).all()

        norm_sem = get_current_cr_semester(db)
        if norm_sem:
            all_courses = db.query(Course).order_by(Course.course_code).all()
            courses = [c for c in all_courses if normalize_semester(c.semester) == norm_sem]
        else:
            courses = []
            flash("আপনার batch এর জন্য কোনো current semester খুঁজে পাওয়া যায়নি (Excel data চেক করুন)।", "error")

    finally:
        db.close()
    return render_template("cr/dashboard.html", cameras=cameras, courses=courses)


@cr_bp.route("/stop-camera", methods=["POST"])
@login_required
@cr_required
def stop_camera():
    db = SessionLocal()
    try:
        camera_id = request.form.get("camera_id")
        if not camera_id:
            flash("Camera সিলেক্ট করুন!", "error")
            return redirect(url_for("cr.dashboard"))

        camera_obj = db.query(Camera).get(camera_id)
        if not camera_obj:
            flash("Camera পাওয়া যায়নি।", "error")
            return redirect(url_for("cr.dashboard"))

        if camera_obj.current_session_id != current_user.session_id:
            flash("এই camera আপনার session চালু করেনি, তাই আপনি এটা বন্ধ করতে পারবেন না।", "error")
            return redirect(url_for("cr.dashboard"))

        stop_command = CameraCommand(
            camera_id=camera_id,
            course_id=None,
            session_id=current_user.session_id,
            status="pending"
        )
        db.add(stop_command)

        camera_obj.current_course_id = None
        camera_obj.current_session_id = None

        db.commit()
        flash("Camera বন্ধ করার command পাঠানো হয়েছে!", "success")
    finally:
        db.close()
    return redirect(url_for("cr.dashboard"))


@cr_bp.route("/course/<int:course_id>/report")
@login_required
@cr_required
def course_report(course_id):
    db = SessionLocal()
    try:
        course_obj = db.query(Course).get(course_id)

        norm_sem = get_current_cr_semester(db)
        if not course_obj or not norm_sem or normalize_semester(course_obj.semester) != norm_sem:
            flash("এই কোর্সটি আপনার semester-এর অন্তর্গত নয়।", "error")
            return redirect(url_for("cr.dashboard"))

        enrolled = db.query(Enrollment).filter(
            Enrollment.course_id == course_id
        ).all()

        all_dates = db.query(Attendance.date).filter(
            Attendance.role == "student",
            Attendance.course_id == course_id
        ).distinct().all()
        total_days = len(all_dates)

        report = []
        for e in enrolled:
            present_count = db.query(Attendance).filter(
                Attendance.course_id == course_id,
                Attendance.roll_number == e.roll_number
            ).count()
            percentage = round((present_count / total_days * 100), 1) if total_days > 0 else 0
            report.append({
                "roll": e.roll_number,
                "name": e.name,
                "present": present_count,
                "total": total_days,
                "percentage": percentage
            })
        report = sorted(report, key=lambda x: roll_batch_sort_key(x["roll"]))
    finally:
        db.close()
    return render_template("cr/course_report.html", course=course_obj, report=report, total_days=total_days)


@cr_bp.route("/course/<int:course_id>/export")
@login_required
@cr_required
def course_export(course_id):
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    import io, pandas as pd
    from flask import send_file

    db = SessionLocal()
    try:
        course_obj = db.query(Course).get(course_id)

        norm_sem = get_current_cr_semester(db)
        if not course_obj or not norm_sem or normalize_semester(course_obj.semester) != norm_sem:
            flash("এই কোর্সটি আপনার semester-এর অন্তর্গত নয়।", "error")
            return redirect(url_for("cr.dashboard"))

        enrolled = db.query(Enrollment).filter(
            Enrollment.course_id == course_id
        ).all()

        all_dates = db.query(Attendance.date).filter(
            Attendance.role == "student",
            Attendance.course_id == course_id
        ).distinct().all()
        total_days = len(all_dates)

        report = []
        for e in enrolled:
            present_count = db.query(Attendance).filter(
                Attendance.course_id == course_id,
                Attendance.roll_number == e.roll_number
            ).count()
            percentage = round((present_count / total_days * 100), 1) if total_days > 0 else 0
            report.append({
                "roll": e.roll_number,
                "name": e.name,
                "present": present_count,
                "total": total_days,
                "percentage": percentage
            })
        report = sorted(report, key=lambda x: roll_batch_sort_key(x["roll"]))

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            wb = writer.book
            ws = wb.create_sheet("Percentage")
            headers = ["Roll", "Name", "Present", "Total", "Percentage"]
            ws.append(headers)

            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            for row_idx, r in enumerate(report, start=2):
                ws.cell(row=row_idx, column=1, value=r["roll"])
                ws.cell(row=row_idx, column=2, value=r["name"])
                ws.cell(row=row_idx, column=3, value=r["present"])
                ws.cell(row=row_idx, column=4, value=r["total"])
                ws.cell(row=row_idx, column=5, value=f'{r["percentage"]}%')

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        output.seek(0)
        filename = f"{course_obj.course_code}_percentage.xlsx"
    finally:
        db.close()
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )