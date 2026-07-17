from flask import Blueprint, render_template, request, send_file, redirect, url_for, flash
from flask_login import login_required, current_user
from database import SessionLocal, Attendance, load_students_from_excel, load_teachers_from_excel, get_admin_from_env
from sqlalchemy import func
from datetime import datetime
from database import SEMESTER_ORDER, normalize_semester, get_batch_semester_map
from database import Session, Course, Enrollment, CameraCommand, CRAccount, generate_cr_credentials
import pandas as pd
import io
import os
import time

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

_DASHBOARD_CACHE = {}
_DASHBOARD_CACHE_TIME = 0
CACHE_TTL = 60

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ["admin", "teacher"]:
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated

DESIGNATION_RANK = {
    "Professor": 1,
    "Associate Professor": 2,
    "Assistant Professor": 3,
    "Lecturer": 4
}

def get_teacher_rank(designation):
    return DESIGNATION_RANK.get(designation, 99)

def roll_batch_sort_key(roll):
    # roll number এর প্রথম ২ ডিজিট batch/session year বহন করে (যেমন 230501 => batch 23)
    # নিজের/নতুন batch আগে (descending), একই batch এর মধ্যে roll ascending
    roll = roll or ""
    try:
        batch = int(roll[:2])
    except:
        batch = 0
    return (-batch, roll)

def clear_dashboard_cache():
    global _DASHBOARD_CACHE, _DASHBOARD_CACHE_TIME
    _DASHBOARD_CACHE = {}
    _DASHBOARD_CACHE_TIME = 0

@admin_bp.route("/session/add", methods=["POST"])
@login_required
@admin_required
def add_session():
    db = SessionLocal()
    try:
        name = request.form.get("session_name", "").strip()
        if not name:
            flash("Session Name অবশ্যই দিতে হবে! (যেমন: 2025-26)", "error")
        else:
            existing = db.query(Session).filter(Session.name == name).first()
            if existing:
                flash(f"'{name}' নামে Session ইতিমধ্যে আছে!", "error")
            else:
                new_session = Session(name=name, is_active=1)
                db.add(new_session)
                db.commit()
                flash(f"Session '{name}' যোগ হয়েছে!", "success")
    finally:
        db.close()
    return redirect(url_for("admin.dashboard"))

@admin_bp.route("/semester/<int:semester_index>/courses")
@login_required
@admin_required
def semester_courses(semester_index):
    if semester_index < 0 or semester_index >= len(SEMESTER_ORDER):
        flash("ভুল Semester!", "error")
        return redirect(url_for("admin.dashboard"))

    semester_label = SEMESTER_ORDER[semester_index]

    db = SessionLocal()
    try:
        courses = db.query(Course).filter(Course.semester == semester_label)\
            .order_by(Course.course_code).all()
        normalized_semester_to_batch = get_batch_semester_map()
        batch_name = normalized_semester_to_batch.get(normalize_semester(semester_label))
        session_obj = None
        if batch_name:
            session_obj = db.query(Session).filter(
                Session.name == batch_name, Session.is_active == 1
            ).first()
    finally:
        db.close()

    return render_template(
        "admin/courses.html",
        semester_label=semester_label,
        semester_index=semester_index,
        courses=courses,
        session=session_obj
    )

@admin_bp.route("/semester/<int:semester_index>/course/add", methods=["POST"])
@login_required
@admin_required
def add_course(semester_index):
    if semester_index < 0 or semester_index >= len(SEMESTER_ORDER):
        flash("ভুল Semester!", "error")
        return redirect(url_for("admin.dashboard"))
    semester_label = SEMESTER_ORDER[semester_index]
    db = SessionLocal()
    try:
        course_code = request.form.get("course_code", "").strip()
        course_name = request.form.get("course_name", "").strip()
        if not course_code or not course_name:
            flash("Course Code এবং Course Name অবশ্যই দিতে হবে!", "error")
        else:
            existing = db.query(Course).filter(
                Course.semester == semester_label,
                Course.course_code == course_code
            ).first()
            if existing:
                flash("এই Course Code ইতিমধ্যে এই Semester-এ আছে!", "error")
            else:
                new_course = Course(
                    semester=semester_label,
                    course_code=course_code,
                    course_name=course_name
                )
                db.add(new_course)
                db.commit()
                flash(f"Course '{course_code}' যোগ হয়েছে!", "success")
    finally:
        db.close()
    return redirect(url_for("admin.semester_courses", semester_index=semester_index))

@admin_bp.route("/course/<int:course_id>/edit", methods=["POST"])
@login_required
@admin_required
def edit_course(course_id):
    db = SessionLocal()
    try:
        course = db.query(Course).get(course_id)
        if not course:
            flash("Course পাওয়া যায়নি!", "error")
            return redirect(url_for("admin.dashboard"))
        semester_label = course.semester
        semester_index = SEMESTER_ORDER.index(semester_label) if semester_label in SEMESTER_ORDER else 0
        course.course_code = request.form.get("course_code", "").strip()
        course.course_name = request.form.get("course_name", "").strip()
        db.commit()
        flash("Course আপডেট হয়েছে!", "success")
    finally:
        db.close()
    return redirect(url_for("admin.semester_courses", semester_index=semester_index))


@admin_bp.route("/course/<int:course_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_course(course_id):
    db = SessionLocal()
    try:
        course = db.query(Course).get(course_id)
        if not course:
            flash("Course পাওয়া যায়নি!", "error")
            return redirect(url_for("admin.dashboard"))
        semester_label = course.semester
        semester_index = SEMESTER_ORDER.index(semester_label) if semester_label in SEMESTER_ORDER else 0
        db.query(Enrollment).filter(Enrollment.course_id == course_id).delete()
        db.query(CameraCommand).filter(CameraCommand.course_id == course_id).delete()
        db.delete(course)
        db.commit()
        flash("Course এবং সংশ্লিষ্ট enrollment মুছে ফেলা হয়েছে!", "success")
    finally:
        db.close()
    return redirect(url_for("admin.semester_courses", semester_index=semester_index))


@admin_bp.route("/unassigned-attendance", methods=["GET", "POST"])
@login_required
@admin_required
def unassigned_attendance():
    db = SessionLocal()
    try:
        if request.method == "POST":
            course_id = request.form.get("course_id")
            selected_dates = request.form.getlist("dates")
            if not course_id or not selected_dates:
                flash("Course এবং অন্তত একটি তারিখ সিলেক্ট করুন!", "error")
            else:
                updated = db.query(Attendance).filter(
                    Attendance.course_id.is_(None),
                    Attendance.date.in_(selected_dates)
                ).update({Attendance.course_id: course_id}, synchronize_session=False)
                db.commit()
                flash(f"{updated} টি রেকর্ড সফলভাবে assign করা হয়েছে!", "success")
        orphan_dates = db.query(
            Attendance.date, func.count(Attendance.id)
        ).filter(Attendance.course_id.is_(None)).group_by(Attendance.date).order_by(Attendance.date.desc()).all()
        all_courses_raw = db.query(Course).order_by(Course.course_code).all()
        all_courses = sorted(
            all_courses_raw,
            key=lambda c: (
                SEMESTER_ORDER.index(c.semester) if c.semester in SEMESTER_ORDER else 999,
                c.course_code
            )
        )
    finally:
        db.close()
    return render_template("admin/unassigned_attendance.html", orphan_dates=orphan_dates, courses=all_courses)

@admin_bp.route("/session/<int:session_id>/course/<int:course_id>/dashboard")
@login_required
@admin_required
def course_dashboard(session_id, course_id):
    global _DASHBOARD_CACHE, _DASHBOARD_CACHE_TIME

    today = __import__("datetime").date.today().strftime("%Y-%m-%d")
    now = time.time()
    cache_key = f"{session_id}_{course_id}_{today}"

    if _DASHBOARD_CACHE and (now - _DASHBOARD_CACHE_TIME) < CACHE_TTL and _DASHBOARD_CACHE.get("cache_key") == cache_key:
        return render_template("admin/dashboard.html", **_DASHBOARD_CACHE)

    db = SessionLocal()
    try:
        course_obj = db.query(Course).get(course_id)
        all_today_records = db.query(Attendance).filter(
            Attendance.date == today,
            Attendance.course_id == course_id
        ).all()

        def student_obj_sort_key(r):
            sec = r.section if r.section else "0-0"
            try:
                first_year = int(sec.split("-")[0])
            except:
                first_year = 0
            return (-first_year, r.roll_number if r.roll_number else "")

        student_records = sorted(
            [r for r in all_today_records if r.role == "student"],
            key=student_obj_sort_key
        )

        teacher_records = sorted(
            [r for r in all_today_records if r.role == "teacher"],
            key=lambda x: get_teacher_rank(x.section)
        )

        today_records = student_records + teacher_records

        # এখন course-এর সব session (batch) মিলিয়ে সব student একসাথে দেখানো হচ্ছে
        enrolled = db.query(Enrollment).filter(
            Enrollment.course_id == course_id
        ).all()
        students_list = sorted(
            [{"roll": e.roll_number, "name": e.name, "section": None} for e in enrolled],
            key=lambda s: roll_batch_sort_key(s.get("roll"))
        )
        total_students = len(students_list)
        today_present = len(student_records)
        today_absent = total_students - today_present
        sections = sorted(list(set([s["section"] for s in students_list if s.get("section")])))

        present_rolls = [r.roll_number for r in student_records]
        absent_students = sorted(
            [s for s in students_list if s.get("roll") not in present_rolls],
            key=lambda s: roll_batch_sort_key(s.get("roll"))
        )

        teachers_list = []
        absent_teachers = []

        all_dates = db.query(Attendance.date).filter(
            Attendance.role == "student",
            Attendance.course_id == course_id
        ).distinct().all()
        total_days = len(all_dates)

        low_attendance = []
        if total_days > 0:
            attendance_results = db.query(Attendance.roll_number, func.count(Attendance.id))\
                .filter(
                    Attendance.role == "student",
                    Attendance.course_id == course_id
                )\
                .group_by(Attendance.roll_number).all()

            attendance_counts = {row[0]: row[1] for row in attendance_results}

            for student in students_list:
                student_roll = student.get("roll")
                present_count = attendance_counts.get(student_roll, 0)
                percentage = round((present_count / total_days * 100), 1)
                if percentage < 70:
                    low_attendance.append({
                        "roll": student_roll,
                        "name": student.get("name"),
                        "section": student.get("section"),
                        "percentage": percentage,
                        "present": present_count,
                        "total": total_days
                    })

        from datetime import datetime as dt, timedelta
        chart_labels = []
        chart_present = []
        chart_absent = []

        start_date = (dt.today() - timedelta(days=6)).strftime("%Y-%m-%d")

        chart_results = db.query(Attendance.date, func.count(Attendance.roll_number.distinct()))\
            .filter(
                Attendance.date >= start_date,
                Attendance.role == "student",
                Attendance.course_id == course_id
            )\
            .group_by(Attendance.date).all()

        chart_counts = {row[0]: row[1] for row in chart_results}

        for i in range(6, -1, -1):
            day = (dt.today() - timedelta(days=i)).strftime("%Y-%m-%d")
            day_label = (dt.today() - timedelta(days=i)).strftime("%d %b")
            present = chart_counts.get(day, 0)
            chart_labels.append(day_label)
            chart_present.append(present)
            chart_absent.append(total_students - present)

        context = dict(
            course=course_obj,
            session_id=session_id,
            course_id=course_id,
            cache_key=cache_key,
            today_records=today_records,
            total_students=total_students,
            today_present=today_present,
            today_absent=today_absent,
            sections=sections,
            today=today,
            absent_students=absent_students,
            absent_teachers=absent_teachers,
            low_attendance=low_attendance,
            chart_labels=chart_labels,
            chart_present=chart_present,
            chart_absent=chart_absent,
            all_registered_students=students_list,
            all_registered_teachers=teachers_list
        )

        _DASHBOARD_CACHE.clear()
        _DASHBOARD_CACHE.update(context)
        _DASHBOARD_CACHE_TIME = now

    finally:
        db.close()
    return render_template("admin/dashboard.html", **_DASHBOARD_CACHE)

SEMESTER_ORDER = [
    "1st Year 1st Semester",
    "1st Year 2nd Semester",
    "2nd Year 1st Semester",
    "2nd Year 2nd Semester",
    "3rd Year 1st Semester",
    "3rd Year 2nd Semester",
    "4th Year 1st Semester",
    "4th Year 2nd Semester",
]

def normalize_semester(text):
    """
    Excel এ কেউ 'Semester' বা ভুলবশত 'Semister' যেভাবেই লিখুক না কেন,
    দুইটাকেই সমান ধরে match করার জন্য এই normalize function।
    """
    if not text:
        return ""
    return text.strip().lower().replace("semister", "semester").replace("  ", " ")

_NORMALIZED_SEMESTER_ORDER = [
    (idx, label, normalize_semester(label)) for idx, label in enumerate(SEMESTER_ORDER)
]

@admin_bp.route("/dashboard")
@login_required
@admin_required
def dashboard():
    db = SessionLocal()
    try:
        normalized_semester_to_batch = get_batch_semester_map()
        all_sessions = db.query(Session).filter(Session.is_active == 1).all()
        session_by_name = {s.name: s for s in all_sessions}
        semester_cards = []
        empty_semesters = []
        for idx, sem_label in enumerate(SEMESTER_ORDER):
            norm_label = normalize_semester(sem_label)
            batch_name = normalized_semester_to_batch.get(norm_label)
            session_obj = session_by_name.get(batch_name) if batch_name else None
            card = {
                "order": idx,
                "semester_index": idx,
                "semester_label": sem_label,
                "session": session_obj,
            }
            if session_obj:
                semester_cards.append(card)
            else:
                empty_semesters.append(card)
        semester_cards.sort(key=lambda c: c["order"])
        empty_semesters.sort(key=lambda c: c["order"])
        ordered_cards = semester_cards + empty_semesters
    finally:
        db.close()
    return render_template("admin/sessions.html", semester_cards=ordered_cards)

@admin_bp.route("/teachers/attendance")
@login_required
@admin_required
def teacher_attendance():
    today = datetime.now().strftime("%Y-%m-%d")
    db = SessionLocal()
    try:
        today_records = db.query(Attendance).filter(
            Attendance.date == today,
            Attendance.role == "teacher"
        ).all()

        teacher_records = sorted(today_records, key=lambda x: get_teacher_rank(x.section))

        all_teachers = load_teachers_from_excel()
        present_names = [r.name for r in teacher_records]

        absent_teachers = sorted(
            [t for t in all_teachers if t["name"] not in present_names],
            key=lambda t: get_teacher_rank(t.get("designation"))
        )

        total_teachers = len(all_teachers)
        today_present = len(teacher_records)
        today_absent = total_teachers - today_present
    finally:
        db.close()

    return render_template("admin/teacher_attendance.html",
        today=today,
        teacher_records=teacher_records,
        absent_teachers=absent_teachers,
        total_teachers=total_teachers,
        today_present=today_present,
        today_absent=today_absent,
        all_registered_teachers=all_teachers
    )

@admin_bp.route("/session/<int:session_id>/course/<int:course_id>/history")
@login_required
@admin_required
def course_history(session_id, course_id):
    db = SessionLocal()
    try:
        course_obj = db.query(Course).get(course_id)
        date_from = request.args.get("date_from", "")
        date_to = request.args.get("date_to", "")
        query = db.query(Attendance).filter(Attendance.course_id == course_id)
        if date_from:
            query = query.filter(Attendance.date >= date_from)
        if date_to:
            query = query.filter(Attendance.date <= date_to)
        records = query.order_by(Attendance.date).all()
        student_records = [r for r in records if r.role == "student"]
        teacher_records = [r for r in records if r.role == "teacher"]
        student_records = sorted(
            student_records,
            key=lambda r: (r.date, roll_batch_sort_key(r.roll_number))
        )
        teacher_records = sorted(
            teacher_records,
            key=lambda r: (r.date, roll_batch_sort_key(r.roll_number))
        )
    finally:
        db.close()
    return render_template("admin/course_history.html",
        course=course_obj,
        session_id=session_id,
        student_records=student_records,
        teacher_records=teacher_records,
        date_from=date_from,
        date_to=date_to
    )


@admin_bp.route("/session/<int:session_id>/course/<int:course_id>/percentage")
@login_required
@admin_required
def course_percentage(session_id, course_id):
    db = SessionLocal()
    try:
        course_obj = db.query(Course).get(course_id)
        date_from = request.args.get("date_from", "")
        date_to = request.args.get("date_to", "")
        enrolled = db.query(Enrollment).filter(Enrollment.course_id == course_id).all()
        days_query = db.query(Attendance.date).filter(
            Attendance.course_id == course_id,
            Attendance.role == "student"
        )
        if date_from:
            days_query = days_query.filter(Attendance.date >= date_from)
        if date_to:
            days_query = days_query.filter(Attendance.date <= date_to)
        total_days = days_query.distinct().count()
        count_query = db.query(Attendance.roll_number, func.count(Attendance.id))\
            .filter(Attendance.course_id == course_id, Attendance.role == "student")
        if date_from:
            count_query = count_query.filter(Attendance.date >= date_from)
        if date_to:
            count_query = count_query.filter(Attendance.date <= date_to)
        attendance_results = count_query.group_by(Attendance.roll_number).all()
        attendance_counts = {row[0]: row[1] for row in attendance_results}
        result = []
        for e in enrolled:
            present_count = attendance_counts.get(e.roll_number, 0)
            percentage = round((present_count / total_days * 100), 1) if total_days > 0 else 0
            result.append({
                "roll": e.roll_number,
                "name": e.name,
                "present": present_count,
                "total": total_days,
                "percentage": percentage
            })
        result = sorted(result, key=lambda x: roll_batch_sort_key(x.get("roll")))
    finally:
        db.close()
    return render_template("admin/course_percentage.html",
        course=course_obj,
        session_id=session_id,
        result=result,
        date_from=date_from,
        date_to=date_to
    )

@admin_bp.route("/session/<int:session_id>/course/<int:course_id>/export/percentage")
@login_required
@admin_required
def course_export_percentage(session_id, course_id):
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    db = SessionLocal()
    try:
        course_obj = db.query(Course).get(course_id)

        date_from = request.args.get("date_from", "")
        date_to = request.args.get("date_to", "")

        enrolled = db.query(Enrollment).filter(Enrollment.course_id == course_id).all()

        days_query = db.query(Attendance.date).filter(
            Attendance.course_id == course_id,
            Attendance.role == "student"
        )
        if date_from:
            days_query = days_query.filter(Attendance.date >= date_from)
        if date_to:
            days_query = days_query.filter(Attendance.date <= date_to)
        total_days = days_query.distinct().count()

        count_query = db.query(Attendance.roll_number, func.count(Attendance.id))\
            .filter(Attendance.course_id == course_id, Attendance.role == "student")
        if date_from:
            count_query = count_query.filter(Attendance.date >= date_from)
        if date_to:
            count_query = count_query.filter(Attendance.date <= date_to)
        attendance_results = count_query.group_by(Attendance.roll_number).all()
        attendance_counts = {row[0]: row[1] for row in attendance_results}

        result = []
        for e in enrolled:
            present_count = attendance_counts.get(e.roll_number, 0)
            percentage = round((present_count / total_days * 100), 1) if total_days > 0 else 0
            result.append({
                "roll": e.roll_number,
                "name": e.name,
                "present": present_count,
                "total": total_days,
                "percentage": percentage
            })
        result = sorted(result, key=lambda x: roll_batch_sort_key(x.get("roll")))

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            wb = writer.book
            ws = wb.create_sheet("Percentage")
            headers = ["Roll", "Name", "Present", "Total", "Percentage"]
            ws.append(headers)

            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            for row_idx, r in enumerate(result, start=2):
                ws.cell(row=row_idx, column=1, value=r["roll"])
                ws.cell(row=row_idx, column=2, value=r["name"])
                ws.cell(row=row_idx, column=3, value=r["present"])
                ws.cell(row=row_idx, column=4, value=r["total"])
                ws.cell(row=row_idx, column=5, value=f'{r["percentage"]}%')

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        output.seek(0)
        filename = f"{course_obj.course_code}_percentage.xlsx"
    finally:
        db.close()

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@admin_bp.route("/session/<int:session_id>/course/<int:course_id>/export/history")
@login_required
@admin_required
def course_export_history(session_id, course_id):
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    db = SessionLocal()
    try:
        course_obj = db.query(Course).get(course_id)

        date_from = request.args.get("date_from", "")
        date_to = request.args.get("date_to", "")

        query = db.query(Attendance).filter(
            Attendance.course_id == course_id,
            Attendance.role == "student"
        )
        if date_from:
            query = query.filter(Attendance.date >= date_from)
        if date_to:
            query = query.filter(Attendance.date <= date_to)

        records = query.order_by(Attendance.date, Attendance.roll_number).all()

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            wb = writer.book
            ws = wb.create_sheet("History")
            headers = ["Date", "Roll", "Name", "Time", "Status"]
            ws.append(headers)

            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            current_date = None
            row_num = 2
            for r in records:
                if r.date != current_date:
                    current_date = r.date
                    date_cell = ws.cell(row=row_num, column=1, value=r.date)
                    date_cell.font = Font(bold=True, size=12, color="000000")
                    date_cell.fill = PatternFill(start_color="E8E4FF", end_color="E8E4FF", fill_type="solid")
                    ws.row_dimensions[row_num].height = 20
                    row_num += 1

                ws.cell(row=row_num, column=1, value="")
                ws.cell(row=row_num, column=2, value=r.roll_number)
                ws.cell(row=row_num, column=3, value=r.name)
                ws.cell(row=row_num, column=4, value=r.time)
                ws.cell(row=row_num, column=5, value=r.status)
                row_num += 1

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        output.seek(0)
        filename = f"{course_obj.course_code}_history.xlsx"
    finally:
        db.close()

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@admin_bp.route("/send-notifications", methods=["POST"])
@login_required
@admin_required
def send_notifications():
    from email_sender import send_absent_notifications
    target_date = request.form.get("date", "")
    course_id = request.form.get("course_id", type=int)      # <-- এই দুই লাইন এখানে বসাও
    session_id = request.form.get("session_id", type=int)    # <--
    if not target_date:
        from datetime import date
        target_date = date.today().strftime("%Y-%m-%d")

    result = send_absent_notifications(target_date)
    
    flash(
        f"{target_date} তারিখের মোট {result['total_absent']} জন অনুপস্থিত — "
        f"{result['success']} জনকে email পাঠানো হয়েছে, "
        f"{result['failed']} জন failed।",
        "success"
    )
    return redirect(url_for('admin.course_dashboard', session_id=session_id, course_id=course_id))

@admin_bp.route("/upload-students", methods=["GET", "POST"])
@login_required
@admin_required
def upload_students():
    db = SessionLocal()
    try:
        semesters = SEMESTER_ORDER

        if request.method == "POST":
            file = request.files.get("excel_file")
            selected_semester = request.form.get("semester")

            if not file or not file.filename.endswith(".xlsx"):
                flash("শুধু .xlsx ফাইল upload করুন!", "error")
                return render_template("admin/upload_students.html", semesters=semesters)

            if not selected_semester:
                flash("কোন Semester এর জন্য sheet, সেটা সিলেক্ট করুন!", "error")
                return render_template("admin/upload_students.html", semesters=semesters)

            import pandas as pd
            df = pd.read_excel(file)

            df = pd.read_excel(save_path)
            df.columns = df.columns.str.strip().str.lower()

            required_cols = {"name", "roll", "session", "semester"}
            if not required_cols.issubset(set(df.columns)):
                flash(
                    f"⚠️ Excel ফাইলে এই কলামগুলো থাকতে হবে: {', '.join(required_cols)}। "
                    f"পাওয়া গেছে: {', '.join(df.columns)}",
                    "error"
                )
                return redirect(url_for("admin.dashboard"))

            courses = db.query(Course).filter(Course.semester == selected_semester).all()
            course_ids = [c.id for c in courses]

            if not course_ids:
                flash(
                    f"'{selected_semester}'-এর জন্য এখনো কোনো course add করা হয়নি। "
                    f"আগে ওই semester-এ course add করুন, তারপর আবার এই ফাইল upload করুন।",
                    "error"
                )
                return redirect(url_for("admin.dashboard"))

            enrolled_count = 0
            updated_count = 0
            skipped_count = 0
            mismatched_rows = 0
            sessions_touched = set()
            session_cache = {}

            for _, row in df.iterrows():
                roll = str(row.get("roll", "")).strip()
                name = str(row.get("name", "")).strip()
                row_session_name = str(row.get("session", "")).strip()
                row_semester_raw = str(row.get("semester", "")).strip()

                if not roll or not row_session_name:
                    continue

                norm_row_semester = normalize_semester(row_semester_raw)
                norm_selected_semester = normalize_semester(selected_semester)
                if norm_row_semester != norm_selected_semester:
                    mismatched_rows += 1
                    continue

                if row_session_name in session_cache:
                    session_obj = session_cache[row_session_name]
                else:
                    session_obj = db.query(Session).filter(Session.name == row_session_name).first()
                    if not session_obj:
                        session_obj = Session(name=row_session_name)
                        db.add(session_obj)
                        db.flush()
# নতুন session-এর জন্য অটোমেটিক CR account তৈরি
                        existing_cr = db.query(CRAccount).filter(
                            CRAccount.session_id == session_obj.id
                        ).first()
                        if not existing_cr:
                            cr_email, cr_password = generate_cr_credentials(session_obj.name)
                            new_cr = CRAccount(
                                session_id=session_obj.id,
                                name=f"CR - {session_obj.name}",
                                login_email=cr_email,
                                login_password=cr_password
                            )
                            db.add(new_cr)

                    session_cache[row_session_name] = session_obj

                sessions_touched.add(row_session_name)

                for cid in course_ids:
                    existing = db.query(Enrollment).filter(
                        Enrollment.course_id == cid,
                        Enrollment.session_id == session_obj.id,
                        Enrollment.user_id == roll
                    ).first()

                    if existing:
                        if existing.name != name:
                            existing.name = name
                            existing.roll_number = roll
                            updated_count += 1
                        else:
                            skipped_count += 1
                        continue

                    new_enrollment = Enrollment(
                        course_id=cid,
                        session_id=session_obj.id,
                        user_id=roll,
                        name=name,
                        roll_number=roll
                    )
                    db.add(new_enrollment)
                    enrolled_count += 1

            db.commit()

            msg = (
                f"'{selected_semester}' এর Student list update হয়েছে! "
                f"{len(sessions_touched)} টি session পাওয়া গেছে ({', '.join(sorted(sessions_touched))}). "
                f"{enrolled_count} নতুন enrollment, {updated_count} নাম আপডেট হয়েছে, {skipped_count} অপরিবর্তিত ছিল।"
            )
            if mismatched_rows:
                msg += f" ⚠️ {mismatched_rows} টি row-এর semester সিলেক্ট করা semester-এর সাথে মেলেনি, সেগুলো বাদ দেওয়া হয়েছে।"

            flash(msg, "success")
            return redirect(url_for("admin.dashboard"))

    finally:
        db.close()
    return render_template("admin/upload_students.html", semesters=semesters)

@admin_bp.route("/change-password", methods=["GET", "POST"])
@login_required
@admin_required
def change_password():
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new_pass = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")

        admin = get_admin_from_env()

        if current != admin["login_password"]:
            flash("বর্তমান password ভুল!", "error")
            return redirect(url_for("admin.change_password"))

        if new_pass != confirm:
            flash("নতুন password দুটো মিলছে না!", "error")
            return redirect(url_for("admin.change_password"))

        if len(new_pass) < 6:
            flash("Password কমপক্ষে ৬ অক্ষরের হতে হবে!", "error")
            return redirect(url_for("admin.change_password"))

        env_path = os.path.normpath(os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "..", ".env"
        ))

        with open(env_path, "r") as f:
            lines = f.readlines()

        with open(env_path, "w") as f:
            for line in lines:
                if line.startswith("ADMIN_PASSWORD="):
                    f.write(f"ADMIN_PASSWORD={new_pass}\n")
                else:
                    f.write(line)

        from dotenv import load_dotenv
        load_dotenv(env_path, override=True)

        flash("Password সফলভাবে পরিবর্তন হয়েছে!", "success")
        return redirect(url_for("admin.dashboard"))
    return render_template("admin/change_password.html")

@admin_bp.route("/manual-attendance", methods=["POST"])
@login_required
@admin_required
def manual_attendance():
    role = request.form.get("role")
    identifier = request.form.get("identifier", "").strip()
    status = request.form.get("status", "On Time")
    custom_date = request.form.get("update_date")
    course_id = request.form.get("course_id", type=int)
    session_id = request.form.get("session_id", type=int)
    if not custom_date:
        custom_date = datetime.now().strftime("%Y-%m-%d")

    time_str = datetime.now().strftime("%I:%M %p")
    db = SessionLocal()

    try:
        if role == "student" and identifier.lower() == "all":
            # course_id দিয়ে filter, session_id দিয়ে না —
            # কারণ একই course-এ multiple session (batch)-এর student enrolled থাকতে পারে
            enrolled_students = db.query(Enrollment).filter(
                Enrollment.course_id == course_id
            ).all()

            course_obj = db.query(Course).filter(Course.id == course_id).first()
            course_semester = course_obj.semester if course_obj else ""

            count = 0
            for enrollment in enrolled_students:
                exists = db.query(Attendance).filter(
                    Attendance.name == enrollment.name,
                    Attendance.date == custom_date,
                    Attendance.course_id == course_id
                ).first()
                if not exists:
                    new_record = Attendance(
                        user_id=enrollment.user_id,
                        name=enrollment.name,
                        role="student",
                        roll_number=enrollment.roll_number,
                        section="",
                        date=custom_date,
                        time=time_str,
                        status=status,
                        semester=course_semester,
                        course_id=course_id,
                        session_id=enrollment.session_id
                    )
                    db.add(new_record)
                    count += 1
            db.commit()
            clear_dashboard_cache()
            flash(f"সফলভাবে মোট {count} জন শিক্ষার্থীর বাল্ক হাজিরা নেওয়া হয়েছে।", "success")

        else:
            target_name = identifier
            target_roll = ""
            target_section = ""
            target_semester = ""
            db_role = "student" if role == "student" else "teacher"

            if role == "student":
                all_students = load_students_from_excel()
                student_match = next(
                    (s for s in all_students if str(s.get("roll")) == identifier), None
                )
                if student_match:
                    target_name = student_match["name"]
                    target_roll = student_match.get("roll", "")
                    target_section = student_match.get("section", "")
                    target_semester = student_match.get("semester", "")
            else:
                all_teachers = load_teachers_from_excel()
                teacher_match = next(
                    (t for t in all_teachers if t["name"] == identifier), None
                )
                if teacher_match:
                    target_name = teacher_match["name"]
                    target_section = teacher_match.get("designation", "")

            exists = db.query(Attendance).filter(
                Attendance.name == target_name,
                Attendance.date == custom_date,
                Attendance.course_id == course_id
            ).first()

            if exists:
                flash(f"দুঃখিত, {target_name} এর হাজিরা আজ আগেই নেওয়া হয়েছে!", "danger")
            else:
                new_record = Attendance(
                    user_id=target_roll if target_roll else target_name,
                    name=target_name,
                    role=db_role,
                    roll_number=target_roll,
                    section=target_section,
                    date=custom_date,
                    time=time_str,
                    status=status,
                    semester=target_semester,
                    course_id=course_id,
                    session_id=session_id
                )
                db.add(new_record)
                db.commit()
                clear_dashboard_cache()
                flash(f"সফলভাবে {target_name} এর ম্যানুয়াল হাজিরা নেওয়া হয়েছে।", "success")

    except Exception as e:
        db.rollback()
        print(f"Error: {e}")
    finally:
        db.close()
    return redirect(url_for('admin.course_dashboard', session_id=session_id, course_id=course_id))    

@admin_bp.route('/update_attendance', methods=['POST'])
def update_attendance():
    date_to_update = request.form.get('update_date')
    role_type = request.form.get('role')
    identifier = request.form.get('student_roll')
    new_status = request.form.get('new_status')
    course_id = request.form.get('course_id', type=int)
    session_id = request.form.get('session_id', type=int)

    if date_to_update and identifier and new_status:
        session = SessionLocal()
        try:
            if role_type == 'student':
                record = session.query(Attendance).filter_by(
                    date=date_to_update, roll_number=identifier, course_id=course_id
                ).first()
            else:
                record = session.query(Attendance).filter_by(
                    date=date_to_update, name=identifier, course_id=course_id
                ).first()

            if record:
                if new_status == 'Absent':
                    session.delete(record)
                    session.commit()
                    clear_dashboard_cache()
                    flash(f"রেকর্ড মুছে ফেলা হয়েছে! এখন তাকে Absent লিস্টে দেখাবে।", "success")
                else:
                    record.status = new_status
                    session.commit()
                    clear_dashboard_cache()
                    flash(f"হাজিরা সফলভাবে '{new_status}' করা হয়েছে!", "success")
            else:
                flash(f"{date_to_update} তারিখে এই ব্যক্তির কোনো হাজিরার রেকর্ড পাওয়া যায়নি।", "warning")

        except Exception as e:
            session.rollback()
            flash("স্ট্যাটাস আপডেট করতে গিয়ে একটি সমস্যা হয়েছে।", "danger")
            print(f"[ERROR] {e}")
        finally:
            session.close()
    else:
        flash("অনুগ্রহ করে তারিখ, ব্যক্তি এবং নতুন স্ট্যাটাস নির্বাচন করুন।", "danger")

    return redirect(url_for('admin.course_dashboard', session_id=session_id, course_id=course_id))

@admin_bp.route('/delete_attendance', methods=['POST'])
def delete_attendance():
    date_to_delete = request.form.get('delete_date')
    delete_scope = request.form.get('delete_scope')
    course_id = request.form.get('course_id', type=int)
    session_id = request.form.get('session_id', type=int)

    if date_to_delete:
        session = SessionLocal()
        try:
            query = session.query(Attendance).filter_by(date=date_to_delete, course_id=course_id)

            if delete_scope == 'student':
                records_to_delete = query.filter_by(role='student').all()
            elif delete_scope == 'teacher':
                records_to_delete = query.filter_by(role='teacher').all()
            else:
                records_to_delete = query.all()

            if records_to_delete:
                for record in records_to_delete:
                    session.delete(record)
                session.commit()
                clear_dashboard_cache()
                flash(f"{date_to_delete} তারিখের সিলেক্টেড রেকর্ড সফলভাবে মুছে ফেলা হয়েছে!", "success")
            else:
                flash(f"{date_to_delete} তারিখে কোনো হাজিরার রেকর্ড পাওয়া যায়নি।", "warning")

        except Exception as e:
            session.rollback()
            flash("রেকর্ড মুছে ফেলতে গিয়ে একটি সমস্যা হয়েছে।", "danger")
            print(f"[ERROR] {e}")
        finally:
            session.close()
    else:
        flash("অনুগ্রহ করে একটি তারিখ নির্বাচন করুন।", "danger")

    return redirect(url_for('admin.course_dashboard', session_id=session_id, course_id=course_id))

@admin_bp.route("/camera-control")
@admin_required
def camera_control():
    from database import get_all_cameras_with_status
    cameras = get_all_cameras_with_status()
    return render_template("admin/camera_control.html", cameras=cameras)